from openpyxl import load_workbook
import os
import glob
import pandas as pd
from datetime import date
from package.extract_data import extract_data # import homemade function: 'from "package name"."python file" import "function"'


####################
### Process Data ###
####################

# specify folder containing data
path = './latest_data'

# get all the .xlsx files in the path folder
file_list = glob.glob(os.path.join(path, '*.xlsx')) 

# define worksheet names
o_traf = 'Overall traffic'
pop_cont = 'Popular content'
use_dev = 'Usage by device'
use_time = 'Usage by time'

working_data_dir = './working_data'
date_today = date.today().strftime('%Y-%m-%d')

# # clear out the working data folder  :::: might be better to keep historic files for longer view ::::

# for file in os.listdir(working_data_dir):
#     file_path = os.path.join(working_data_dir,file)
#     if os.path.isfile(file_path) or os.path.islink(file_path): # remove files and symbolic links
#         os.unlink(file_path)

### EXTRACT ALL AGGREGATE DATA FROM THE OVERALL TRAFFIC WORKSHEET ###

def extract_all_agg_traffic(file):

    workbook = load_workbook(filename=file, read_only=True, data_only=True) # open workbook
    worksheet = workbook[o_traf] # select specified worksheet

    agg_traffic = extract_data(worksheet,'A8','C12') # data range within worksheet
    agg_traffic.columns = agg_traffic.iloc[0] # promote first row to headers
    agg_traffic = agg_traffic[1:].reset_index(drop=True) # drop the row that previously contained the headers
    agg_traffic['ICB'] = file.rsplit("\\", 1)[-1].split("-", 1)[0] # get the ICB from the file name. Needs double backslash since it's normally an escape character
    agg_traffic.replace("Not supported","-", inplace=True) # get rid of any instances of "not supported"

    # this next bit doesn't seem to be working.
    agg_traffic[['Unique viewers','Site visits']] = agg_traffic[['Unique viewers','Site visits']].apply(lambda x: int(float(x)) if str(x).replace('.', '', 1).isdigit() else x).fillna(agg_traffic[['Unique viewers','Site visits']])
    

    return agg_traffic

all_agg_traffic = pd.concat([extract_all_agg_traffic(file) for file in file_list], ignore_index=True)

all_agg_traffic.to_csv(f'{working_data_dir}/all_agg_traffic_{date_today}.csv')

### EXTRACT ALL TRAFFIC IN LAST 90 DAYS FROM THE OVERALL TRAFFIC WORKSHEET ###

def extract_all_last90_traffic(file):

    workbook = load_workbook(filename=file, read_only=True, data_only=True) # open workbook
    worksheet = workbook[o_traf] # select specified worksheet
    
    last90_traffic = extract_data(worksheet,'A16','C105') # data range within worksheet
    last90_traffic.columns = last90_traffic.iloc[0] # promote first row to headers
    last90_traffic = last90_traffic[1:].reset_index(drop=True) # drop the row that previously contained the headers
    last90_traffic['ICB'] = file.rsplit("\\", 1)[-1].split("-", 1)[0] # get the ICB from the file name. Needs double backslash since it's normally an escape character
    
    return last90_traffic

all_last90_traffic = pd.concat([extract_all_last90_traffic(file) for file in file_list], ignore_index=True)

all_last90_traffic.to_csv(f'{working_data_dir}/all_last90_traffic_{date_today}.csv')

### EXTRACT ALL HIGH TRAFFIC CONTENT IN LAST 7 DAYS FROM THE POPULAR CONTENT WORKSHEET ###

def extract_all_popular_content(file):

    workbook = load_workbook(filename=file, read_only=True, data_only=True) # open workbook
    worksheet = workbook[pop_cont] # select specified worksheet

    # slightly different since the data range is variable depending on the number of popular reports.

    start_row = 6
    start_col = 'A'
    end_col = 'D'

    end_row = start_row
    for row in range(start_row, worksheet.max_row + 1):
        if all(worksheet[f'{col}{row}'].value is None for col in 'ABCD'):
            break
        end_row = row

    popular_content = extract_data(worksheet,f'{start_col}{start_row}',f'{end_col}{end_row}') # extract from dynamic data range
    popular_content.columns = popular_content.iloc[0] # promote first row to headers
    popular_content = popular_content[1:].reset_index(drop=True) # drop the row that previously contained the headers
    popular_content['ICB'] = file.rsplit("\\", 1)[-1].split("-", 1)[0] # get the ICB from the file name. Needs double backslash since it's normally an escape character

    return popular_content

all_popular_content = pd.concat([extract_all_popular_content(file) for file in file_list], ignore_index=True)

all_popular_content.to_csv(f'{working_data_dir}/all_popular_content_{date_today}.csv')

### EXTRACT ALL USAGE BY DEVICE FROM THE USAGE BY DEVICE WORKSHEET ###

def extract_all_usage_by_device(file):

    workbook = load_workbook(filename=file, read_only=True, data_only=True) # open workbook
    worksheet = workbook[use_dev] # select specified worksheet
    
    usage_by_device = extract_data(worksheet,'A6','F95') # data range within worksheet
    usage_by_device.columns = usage_by_device.iloc[0] # promote first row to headers
    usage_by_device = usage_by_device[1:].reset_index(drop=True) # drop the row that previously contained the headers
    usage_by_device['ICB'] = file.rsplit("\\", 1)[-1].split("-", 1)[0] # get the ICB from the file name. Needs double backslash since it's normally an escape character
    
    return usage_by_device

all_usage_by_device = pd.concat([extract_all_usage_by_device(file) for file in file_list], ignore_index=True)

### EXTRACT ALL USAGE BY TIME FROM THE USAGE BY TIME WORKSHEET ###

def extract_all_usage_by_time(file):

    workbook = load_workbook(filename=file, read_only=True, data_only=True) # open workbook
    worksheet = workbook[use_time] # select specified worksheet
    
    usage_by_time = extract_data(worksheet,'A7','D175') # data range within worksheet
    usage_by_time.columns = usage_by_time.iloc[0] # promote first row to headers
    usage_by_time = usage_by_time[1:].reset_index(drop=True) # drop the row that previously contained the headers
    usage_by_time['ICB'] = file.rsplit("\\", 1)[-1].split("-", 1)[0] # get the ICB from the file name. Needs double backslash since it's normally an escape character
    
    return usage_by_time

all_usage_by_time = pd.concat([extract_all_usage_by_time(file) for file in file_list], ignore_index=True)

all_usage_by_time.to_csv(f'{working_data_dir}/all_usage_by_time_{date_today}.csv')