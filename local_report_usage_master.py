### TODO: ###
#############

'''When I come to set this up in production, I ought to pull in extract_data.py
but this will require me to set up the folder without a leading numeric value in the name,
because you can't use module names starting with a numeric value.
'''

from openpyxl import load_workbook
import os
import glob
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import numpy as np
import altair as alt
from package.extract_data import extract_data # import homemade function: 'from "package name"."python file" import "function"'

# set seaborn theme

sns.set_theme()

# specify folder containing data
path = './latest_data'

# get all the .xlsx files in the folder
file_list = glob.glob(os.path.join(path, '*.xlsx')) 

# define worksheet names
o_traf = 'Overall traffic'
pop_cont = 'Popular content'
use_dev = 'Usage by device'
use_time = 'Usage by time'

### EXTRACT ALL AGGREGATE DATA FROM THE OVERALL TRAFFIC WORKSHEET ###

def extract_all_agg_traffic(file):

    workbook = load_workbook(filename=file, read_only=True, data_only=True) # open workbook
    worksheet = workbook[o_traf] # select specified worksheet

    agg_traffic = extract_data(worksheet,'A8','C12') # data range within worksheet
    agg_traffic.columns = agg_traffic.iloc[0] # promote first row to headers
    agg_traffic = agg_traffic[1:].reset_index(drop=True) # drop the row that previously contained the headers
    agg_traffic['ICB'] = file.rsplit("\\", 1)[-1].split("-", 1)[0] # get the ICB from the file name. Needs double backslash since it's normally an escape character
    agg_traffic.replace("Not supported","-", inplace=True) # get rid of any instances of "not supported"

    # this next bit doesn't seem to be working.
    agg_traffic[['Unique viewers','Site visits']] = agg_traffic[['Unique viewers','Site visits']].apply(lambda x: int(float(x)) if str(x).replace('.', '', 1).isdigit() else x).fillna(agg_traffic[['Unique viewers','Site visits']])
    

    return agg_traffic

all_agg_traffic = pd.concat([extract_all_agg_traffic(file) for file in file_list], ignore_index=True)

### EXTRACT ALL TRAFFIC IN LAST 90 DAYS FROM THE OVERALL TRAFFIC WORKSHEET ###

def extract_all_last90_traffic(file):

    workbook = load_workbook(filename=file, read_only=True, data_only=True) # open workbook
    worksheet = workbook[o_traf] # select specified worksheet
    
    last90_traffic = extract_data(worksheet,'A16','C105') # data range within worksheet
    last90_traffic.columns = last90_traffic.iloc[0] # promote first row to headers
    last90_traffic = last90_traffic[1:].reset_index(drop=True) # drop the row that previously contained the headers
    last90_traffic['ICB'] = file.rsplit("\\", 1)[-1].split("-", 1)[0] # get the ICB from the file name. Needs double backslash since it's normally an escape character
    
    return last90_traffic

all_last90_traffic = pd.concat([extract_all_last90_traffic(file) for file in file_list], ignore_index=True)

### EXTRACT ALL HIGH TRAFFIC CONTENT IN LAST 7 DAYS FROM THE POPULAR CONTENT WORKSHEET ###

def extract_all_popular_content(file):

    workbook = load_workbook(filename=file, read_only=True, data_only=True) # open workbook
    worksheet = workbook[pop_cont] # select specified worksheet

    # slightly different since the data range is variable depending on the number of popular reports.

    start_row = 6
    start_col = 'A'
    end_col = 'D'

    end_row = start_row
    for row in range(start_row, worksheet.max_row + 1):
        if all(worksheet[f'{col}{row}'].value is None for col in 'ABCD'):
            break
        end_row = row

    popular_content = extract_data(worksheet,f'{start_col}{start_row}',f'{end_col}{end_row}') # extract from dynamic data range
    popular_content.columns = popular_content.iloc[0] # promote first row to headers
    popular_content = popular_content[1:].reset_index(drop=True) # drop the row that previously contained the headers
    popular_content['ICB'] = file.rsplit("\\", 1)[-1].split("-", 1)[0] # get the ICB from the file name. Needs double backslash since it's normally an escape character

    return popular_content

all_popular_content = pd.concat([extract_all_popular_content(file) for file in file_list], ignore_index=True)

### EXTRACT ALL USAGE BY DEVICE FROM THE USAGE BY DEVICE WORKSHEET ###

def extract_all_usage_by_device(file):

    workbook = load_workbook(filename=file, read_only=True, data_only=True) # open workbook
    worksheet = workbook[use_dev] # select specified worksheet
    
    usage_by_device = extract_data(worksheet,'A6','F95') # data range within worksheet
    usage_by_device.columns = usage_by_device.iloc[0] # promote first row to headers
    usage_by_device = usage_by_device[1:].reset_index(drop=True) # drop the row that previously contained the headers
    usage_by_device['ICB'] = file.rsplit("\\", 1)[-1].split("-", 1)[0] # get the ICB from the file name. Needs double backslash since it's normally an escape character
    
    return usage_by_device

all_usage_by_device = pd.concat([extract_all_usage_by_device(file) for file in file_list], ignore_index=True)

### EXTRACT ALL USAGE BY TIME FROM THE USAGE BY TIME WORKSHEET ###

def extract_all_usage_by_time(file):

    workbook = load_workbook(filename=file, read_only=True, data_only=True) # open workbook
    worksheet = workbook[use_time] # select specified worksheet
    
    usage_by_time = extract_data(worksheet,'A7','D175') # data range within worksheet
    usage_by_time.columns = usage_by_time.iloc[0] # promote first row to headers
    usage_by_time = usage_by_time[1:].reset_index(drop=True) # drop the row that previously contained the headers
    usage_by_time['ICB'] = file.rsplit("\\", 1)[-1].split("-", 1)[0] # get the ICB from the file name. Needs double backslash since it's normally an escape character
    
    return usage_by_time

all_usage_by_time = pd.concat([extract_all_usage_by_time(file) for file in file_list], ignore_index=True)

### GENERATE AGGREGATED TRAFFIC TABLES FOR EACH ICB ###

def agg_traffic_table(icb):
    df = all_agg_traffic[all_agg_traffic['ICB'] == icb].copy()
    # df.fillna('-', inplace=True)
    

    fig, ax = plt.subplots(figsize=(8,4))
    ax.axis('off')

    table = ax.table(cellText=df.values, colLabels=df.columns, loc='center', cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(12)
    table.auto_set_column_width(col=list(range(len(df.columns))))
    table.scale(1.2,1.2)

    for (row, col), cell in table.get_celld().items():
        if row == 0:  # Header row
            cell.set_facecolor('#1C355E') # SCW dark blue
            cell.set_text_props(color='#FFFFFF') # white

    fig.tight_layout()

    directory = './output'
    filename = f'{icb}_aggregated_overall_traffic.png'
    plt.savefig(f'{directory}/{filename}', bbox_inches='tight', pad_inches=0.01)

    plt.close()


agg_traffic_table('BOB')
agg_traffic_table('Frimley')
agg_traffic_table('HIOW')
agg_traffic_table('Somerset')
agg_traffic_table('Sussex')

import matplotlib.dates as mdates

def last90_chart(icb):
    # Filter the data for the given icb
    df = all_last90_traffic[all_last90_traffic['ICB'] == icb].copy()
    
    # Convert the 'Date' column to datetime explicitly
    df['Date'] = pd.to_datetime(df['Date'])
    
    # Set the 'Date' column as the index to make it a DatetimeIndex
    df.set_index('Date', inplace=True)

    # Set up the figure and axes
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 10)) #, sharex=True

    # Bar plot for Unique viewers (Top subplot)
    ax1.bar(                                     
        df.index, 
        df['Unique viewers'], 
        color='#1C355E', 
        label='Unique viewers',
        alpha=0.8
    )
    ax1.set_ylabel('Unique viewers', color='#1C355E')
    ax1.tick_params(axis='y', labelcolor='#1C355E')
    ax1.set_title(f'Site visits and unique viewers for {icb} ICB')

    ax1.xaxis.set_major_locator(mdates.DayLocator())  # Set major locator for each day
    ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))  # Format dates
    ax1.set_xticks(df.index[::7])  # Set x-ticks to every 7th date
    ax1.set_xticklabels(df.index[::7].strftime('%Y-%m-%d'), rotation=90)

    # Line plot for Site visits (Bottom subplot)
    ax2.bar(
        df.index, 
        df['Site visits'], 
        color='#AE2573', 
        label='Site visits', 
        alpha=0.8
    )
    ax2.set_ylabel('Site visits', color='#AE2573')
    ax2.tick_params(axis='y', labelcolor='#AE2573')

    ax2.xaxis.set_major_locator(mdates.DayLocator())  # Set major locator for each day
    ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))  # Format dates
    ax2.set_xticks(df.index[::7])  # Set x-ticks to every 7th date
    ax2.set_xticklabels(df.index[::7].strftime('%Y-%m-%d'), rotation=90)

    # Adjust layout to prevent overlap
    plt.tight_layout()

    # Save the figure
    directory = './output'
    filename = f'{icb}_last90_activity.png'
    plt.savefig(f'{directory}/{filename}', bbox_inches='tight', pad_inches=0.01)

    # Close the figure to free memory
    plt.close()

last90_chart('BOB')
last90_chart('Frimley')
last90_chart('HIOW')
last90_chart('Somerset')
last90_chart('Sussex')

### CREATE PRETTY POPULAR CONTENT TABLE ###

from matplotlib.colors import Normalize, to_rgba
from matplotlib.cm import ScalarMappable

# Calculate brightness of a color based on its RGBA values for dynamically changing font colour

def calculate_brightness(rgba_color):
    r, g, b, _ = rgba_color  # Extract RGB components
    brightness = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return brightness

# Generate Popular Content Tables

def pop_cont_table(icb):
    df = all_popular_content[all_popular_content['ICB'] == icb].copy()
    df.fillna('-', inplace=True)
    df.rename(columns={'Type (Click to view)': 'Type'}, inplace=True)

    # Define columns for conditional formatting
    columns_to_format = ["Last 7 days unique viewers", "Last 7 days visits"]

    # Normalize values for color gradient
    norm = Normalize(vmin=df[columns_to_format].min().min(), vmax=df[columns_to_format].max().max())
    cmap = plt.cm.YlGnBu  # Choose a colormap

     # Calculate figure size based on DataFrame size
    num_rows, num_columns = df.shape
    fig_width = max(8, num_columns * 1.2)  # Minimum width of 8, scale with columns
    fig_height = max(4, num_rows * 0.25)   # Minimum height of 4, scale with rows

    # Create figure and axis for table
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    ax.axis('off')

    # Create the table
    table = ax.table(cellText=df.values, colLabels=df.columns, loc='center', cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(12)
    table.auto_set_column_width(col=list(range(len(df.columns))))
    table.scale(1.2, 1.2)

    # Apply colour gradient to specific columns
    for (row, col), cell in table.get_celld().items():
        if row > 0:  # Skip the header row
            column_name = df.columns[col]
            if column_name in columns_to_format:
                try:
                    # Get the value and map it to the coloyr range
                    value = df.iloc[row - 1, col]
                    if isinstance(value, (int, float)):  # Ensure value is numeric
                        color = cmap(norm(value))  # Get RGBA colour
                        cell.set_facecolor(color)

                        # Adjust text colour based on brightness
                        brightness = calculate_brightness(to_rgba(color))
                        text_color = '#FFFFFF' if brightness < 0.5 else '#000000'
                        cell.set_text_props(color=text_color)
                except Exception:
                    pass  # Handle non-numeric cells gracefully

    # Customize header row
    for (row, col), cell in table.get_celld().items():
        if row == 0:  # Header row
            cell.set_facecolor('#1C355E')  # SCW dark blue
            cell.set_text_props(color='#FFFFFF')  # White text

    plt.suptitle('Popular Content')

    # Save the table as an image
    directory = './output'
    filename = f'{icb}_popular_content.png'
    plt.savefig(f'{directory}/{filename}', bbox_inches='tight', pad_inches=0.05)

    plt.close()

pop_cont_table('BOB')
pop_cont_table('Frimley')
pop_cont_table('HIOW')
pop_cont_table('Somerset')
pop_cont_table('Sussex')